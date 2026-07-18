"""
Excel Formatter V0.1

Aplica formato visual automático al Excel operativo
de CorrectScoreLab V1.

Objetivos:
- Mantener un diseño consistente en cada ejecución.
- Mejorar la lectura del Dashboard.
- Formatear porcentajes y valores numéricos.
- Ajustar anchos de columnas.
- Aplicar filtros y congelar paneles.
- Diferenciar resultados positivos y negativos.

El formato se aplica DESPUÉS de que Excel Output Engine
haya generado todos los datos.
"""

from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# ==========================================================
# COLORES
# ==========================================================

DARK_BLUE = "1F4E78"
MEDIUM_BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"

DARK_GRAY = "404040"
LIGHT_GRAY = "E7E6E6"

WHITE = "FFFFFF"

GREEN = "C6EFCE"
GREEN_FONT = "006100"

RED = "FFC7CE"
RED_FONT = "9C0006"


# ==========================================================
# ESTILOS BASE
# ==========================================================

thin_border = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def _style_header(row):

    for cell in row:

        if cell.value is None:
            continue

        cell.fill = PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )

        cell.font = Font(
            color=WHITE,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border


def _auto_width(ws, max_width=35):

    for column_cells in ws.columns:

        column_letter = get_column_letter(
            column_cells[0].column
        )

        max_length = 0

        for cell in column_cells:

            if cell.value is None:
                continue

            value = str(cell.value)

            max_length = max(
                max_length,
                len(value)
            )

        adjusted_width = min(
            max_length + 2,
            max_width
        )

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width


def _find_headers(ws):

    header_rows = []

    known_headers = {
        "METRICA",
        "ENGINE",
        "MES",
        "ID_PARTIDO",
        "ID_APUESTA",
        "FECHA",
        "SEASON",
        "PARAMETRO",
    }

    for row in ws.iter_rows():

        values = {
            str(cell.value).strip()
            for cell in row
            if cell.value is not None
        }

        if values.intersection(
            known_headers
        ):

            header_rows.append(
                row[0].row
            )

    return header_rows


def _apply_table_format(ws):

    header_rows = _find_headers(ws)

    for row_number in header_rows:

        _style_header(
            ws[row_number]
        )

        ws.row_dimensions[
            row_number
        ].height = 22


def _format_dates(ws):

    for row in ws.iter_rows():

        for cell in row:

            if (
                cell.column == 1
                and
                cell.row > 1
                and
                hasattr(
                    cell.value,
                    "year"
                )
            ):

                cell.number_format = (
                    "dd/mm/yyyy"
                )


# ==========================================================
# DASHBOARD
# ==========================================================

def _format_dashboard(ws):

    ws.sheet_view.showGridLines = False

    ws.freeze_panes = "A2"

    # ------------------------------------------
    # Bloque principal
    # ------------------------------------------

    _style_header(
        ws[1]
    )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    # ------------------------------------------
    # Métricas principales
    # ------------------------------------------

    for row in range(
        2,
        ws.max_row + 1
    ):

        metric = ws.cell(
            row=row,
            column=1
        ).value

        value_cell = ws.cell(
            row=row,
            column=2
        )

        if metric is None:
            continue

        metric_text = str(
            metric
        ).strip()

        if metric_text in [
            "HIT RATE APUESTAS",
            "ROI",
        ]:

            value_cell.number_format = (
                "0.00%"
            )

        elif metric_text in [
            "STAKE TOTAL",
            "RETORNO",
            "BENEFICIO",
            "BANK INICIAL",
            "BANK FINAL",
            "BANK PEAK",
            "MAX DD PORTFOLIO",
            "MAX DD APUESTAS",
        ]:

            value_cell.number_format = (
                "0.00"
            )

        value_cell.font = Font(
            bold=True
        )

    # ------------------------------------------
    # Encontrar bloques secundarios
    # ------------------------------------------

    engine_header = None
    month_header = None

    for row in range(
        1,
        ws.max_row + 1
    ):

        value = ws.cell(
            row=row,
            column=1
        ).value

        if value == "ENGINE":

            engine_header = row

        elif value == "MES":

            month_header = row

    # ------------------------------------------
    # Motores
    #
    # Performance Engine almacena HIT_RATE
    # y ROI en escala 0-100.
    # Por eso usamos formato con % literal,
    # no formato porcentual Excel.
    # ------------------------------------------

    if engine_header:

        _style_header(
            ws[engine_header]
        )

        for row in range(
            engine_header + 1,
            (
                month_header
                if month_header
                else ws.max_row + 1
            )
        ):

            if (
                ws.cell(
                    row=row,
                    column=1
                ).value
                is None
            ):
                continue

            # HIT_RATE = columna F
            ws.cell(
                row=row,
                column=6
            ).number_format = (
                '0.00"%"'
            )

            # STAKE / RETURN / PROFIT
            for column in [
                7,
                8,
                9,
            ]:

                ws.cell(
                    row=row,
                    column=column
                ).number_format = (
                    "0.00"
                )

            # ROI = columna J
            ws.cell(
                row=row,
                column=10
            ).number_format = (
                '0.00"%"'
            )

    # ------------------------------------------
    # Rendimiento mensual
    # ------------------------------------------

    if month_header:

        _style_header(
            ws[month_header]
        )

        for row in range(
            month_header + 1,
            ws.max_row + 1
        ):

            if (
                ws.cell(
                    row=row,
                    column=1
                ).value
                is None
            ):
                continue

            # HIT_RATE
            ws.cell(
                row=row,
                column=6
            ).number_format = (
                '0.00"%"'
            )

            # STAKE / RETURN / PROFIT
            for column in [
                7,
                8,
                9,
            ]:

                ws.cell(
                    row=row,
                    column=column
                ).number_format = (
                    "0.00"
                )

            # ROI
            ws.cell(
                row=row,
                column=10
            ).number_format = (
                '0.00"%"'
            )

    # ------------------------------------------
    # Formato condicional PROFIT
    # Bloques motores y mensual
    # ------------------------------------------

    if engine_header:

        engine_end = (
            month_header - 1
            if month_header
            else ws.max_row
        )

        profit_range = (
            f"I{engine_header + 1}:"
            f"I{engine_end}"
        )

        ws.conditional_formatting.add(
            profit_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill(
                    "solid",
                    fgColor=GREEN
                ),
                font=Font(
                    color=GREEN_FONT
                )
            )
        )

        ws.conditional_formatting.add(
            profit_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill(
                    "solid",
                    fgColor=RED
                ),
                font=Font(
                    color=RED_FONT
                )
            )
        )

    if month_header:

        profit_range = (
            f"I{month_header + 1}:"
            f"I{ws.max_row}"
        )

        ws.conditional_formatting.add(
            profit_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill(
                    "solid",
                    fgColor=GREEN
                ),
                font=Font(
                    color=GREEN_FONT
                )
            )
        )

        ws.conditional_formatting.add(
            profit_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill(
                    "solid",
                    fgColor=RED
                ),
                font=Font(
                    color=RED_FONT
                )
            )
        )

    _auto_width(
        ws,
        max_width=28
    )


# ==========================================================
# PESTAÑAS OPERATIVAS
# ==========================================================

def _format_operational_sheet(ws):

    ws.sheet_view.showGridLines = False

    ws.freeze_panes = "A2"

    _apply_table_format(
        ws
    )

    _auto_width(
        ws
    )

    if ws.max_row > 1:

        ws.auto_filter.ref = (
            ws.dimensions
        )


# ==========================================================
# FUNCIÓN PRINCIPAL
# ==========================================================

def format_excel(
    excel_file
):

    print()
    print("=" * 100)
    print("EXCEL FORMATTER V0.1")
    print("=" * 100)

    workbook = load_workbook(
        excel_file
    )

    for sheet_name in (
        workbook.sheetnames
    ):

        ws = workbook[
            sheet_name
        ]

        if sheet_name == "DASHBOARD":

            _format_dashboard(
                ws
            )

        else:

            _format_operational_sheet(
                ws
            )

    workbook.save(
        excel_file
    )

    print()
    print(
        "Formato visual aplicado correctamente."
    )

    print()
    print(
        f"Archivo: {excel_file}"
    )

    return excel_file